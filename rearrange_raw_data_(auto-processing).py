import openpyxl as px
import os
os.chdir("path")

wb = px.load_workbook("name.xlsx")
ws_names = wb.sheetnames
ws_num = len(ws_names)
wss=[]
for i in range(ws_num):
    wss.append(wb[ws_names[i]])
#create a new sheet
new_wb = px.Workbook()
#create a new workbook
for i in range(ws_num):
    newsheet = new_wb.create_sheet(ws_names[i]+"-rearranged")
    sample_name = True
    sample_num=0
    while sample_name:
        sample_name = wss[i].cell(row = sample_num+8, column = 2 ).value
        newsheet.cell(row = sample_num+5, column = 1, value=sample_name)
        sample_num += 1
    #enter sample names on the first column in the new sheet
    sample_num -= 1
    compound_num = 0
    compound_name = wss[i].cell(row =  5, column = 1 ).value
    while compound_name:
        newsheet.cell(row = 1, column = compound_num+5, value=compound_name[13:])
        #「enter compound names without「Compound n: 」
        for j in range(sample_num):
            sample_value = wss[i].cell(row = compound_num*(sample_num+4) + 8+j, column = 3 ).value
            if sample_value:
                newsheet.cell(row = j+5, column = compound_num+5, value=sample_value)
            else:
                newsheet.cell(row = j+5, column = compound_num+5, value=0)
        #enter the peakareas of each compound（if the cell is blank, enter "0"）
        compound_num+=1
        compound_name = wss[i].cell(row = compound_num*(sample_num+4) + 5, column = 1 ).value       

    newsheet.cell(row = 2, column = 1, value = "STD?")
    newsheet.cell(row = 3, column = 1, value = "STD group")
    newsheet.cell(row = 4, column = 1, value = "STD quantity(pmol)")
    newsheet.cell(row = 1, column = 2, value = "Blank?")
    newsheet.cell(row = 1, column = 3, value = "Blank group")
    newsheet.cell(row = 1, column = 4, value = "tissue weight(mg)")
    for i in range(2, 5):
        for j in range(2, 4):
            newsheet.cell(row = i, column = j, value = "x")
    for i in range(2, 5):
        newsheet.cell(row = i, column = 4, value = "weight")

new_wb.remove(new_wb["Sheet"])
new_wb.save("name_rearranged.xlsx")