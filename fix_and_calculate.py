import openpyxl as px
import statistics as stx
import os
os.chdir("path")

import process library as mal

wb = px.load_workbook("name_rearranged.xlsx")
ws_names = wb.sheetnames
ws_num = len(ws_names)
wss=[]
for i in range(ws_num):
    wss.append(wb[ws_names[i]])
#create a worksheet list
new_wb = px.Workbook()
#create a new workbook
for i in range(ws_num):
    sample_list = mal.get_sample_list(wss[i])
    if sample_list ==[["sample name"]]:
        continue
    newsheet = new_wb.create_sheet(ws_names[i]+"-processed")
    mal.blank_subtraction(sample_list)
    mal.STD_fix(sample_list)
    mal.tissue_weight_fix(sample_list)
    mal.rearrange(sample_list)
    gathered_sample_list = mal.gather_data(sample_list)
    average_list = mal.cal_average(gathered_sample_list)
    SD_list = mal.cal_SD(gathered_sample_list)
    mal.write_data(0,0,sample_list,newsheet, "pmol/mg tissue")
    lensmpx = len(sample_list)
    lenavex = len(average_list)
    mal.write_data(lensmpx + 2, 0, average_list, newsheet, "pmol/mg tissue")
    mal.write_data(lensmpx + lenavex + 4,0, SD_list, newsheet, "pmol/mg tissue") 

new_wb.remove(new_wb["Sheet"])
new_wb.save("name_processed.xlsx")