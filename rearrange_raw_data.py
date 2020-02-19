import openpyxl as px

wb = px.load_workbook("path")
ws_names = wb.sheetnames
ws_num = len(ws_names)
wss=[]
for i in range(ws_num):
    wss.append(wb[ws_names[i]])
#create a worksheet list
for i in range(ws_num):
    newsheet = wb.create_sheet(ws_names[i]+"-rearranged")
    sample_name = True
    sample_num=0
    while sample_name:
        sample_name = wss[i].cell(row = sample_num+8, column = 2 ).value
        newsheet.cell(row = sample_num+2, column = 1, value=sample_name)
        sample_num += 1
    #enter sample names on the first column in the new sheet
    sample_num -= 1
    compound_num = 0
    compound_name = wss[i].cell(row =  5, column = 1 ).value
    while compound_name:
        newsheet.cell(row = 1, column = compound_num+2, value=compound_name[13:])
        #enter compound names without「Compound n: 」
        for j in range(sample_num):
            sample_value = wss[i].cell(row = compound_num*(sample_num+4) + 8+j, column = 3 ).value
            if sample_value:
                newsheet.cell(row = j+2, column = compound_num+2, value=sample_value)
            else:
                newsheet.cell(row = j+2, column = compound_num+2, value=0)
        #enter the peakareas of each compound（if the cell is blank, enter "0"）
        compound_num+=1
        compound_name = wss[i].cell(row = compound_num*(sample_num+4) + 5, column = 1 ).value       

wb.save("path")