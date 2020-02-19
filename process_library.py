import openpyxl as px
import statistics as stx

def get_sample_list(ws):
    sample_list = [["sample name"]]
    compound_num = 0
    while ws.cell(row = 1 , column= compound_num + 2).value:
        sample_list[0].append(ws.cell(row = 1, column= compound_num + 2).value)
        compound_num += 1
    #get the row for compound name
    sample_num = 0
    while ws.cell(row = sample_num + 2, column = 1).value:
        sample_list.append([])
        for i in range(compound_num + 1):
            sample_list[sample_num + 1].append(ws.cell(row = sample_num + 2, column = i + 1).value)
        sample_num += 1
    #get values by samples
    for i in range(4, sample_num+1):
        map(float,sample_list[i][3:])
    return sample_list

def reverse_2Dlist(either_list):
    temp_list=[]
    for i in range(len(either_list[0])):
        temp_list.append([])
        for j in range(len(either_list)):
            temp_list[i].append(either_list[j][i])
    return temp_list

def blank_subtraction(sample_list):
    sample_num = len(sample_list)
    compound_num = len(sample_list[0])
    blank_dic = {}
    blank_list=[]
    for i in range(4,sample_num):
        if sample_list[i][1] == "blank" or sample_list[i][1] == "Blank" or sample_list[i][1] == "Blank?":
            blank_dic[sample_list[i][2]] = sample_list[i][4:]
            blank_list.append(i)
    #creat a dictionary having blanks
    for i in range(4, compound_num):
        if sample_list[1][i] == "STD" or sample_list[0][i] == "standard":
            for j in blank_list:
                blank_dic[sample_list[j][2]][i-4] = 0
    #set std value of std as 0
    for i in range(4,sample_num):
        blank_group = sample_list[i][2]
        if blank_group in blank_dic:
            for j in range(4,compound_num):
                sample_list[i][j] -= blank_dic[blank_group][j-4]
                if sample_list[i][j] < 0:
                    sample_list[i][j] = 0
        else:
            sample_list[i][2] = "not fixed"
    #fix by blank (if a value become less than 0, enter 0)
    return
        
def STD_fix(sample_list):
    sample_num = len(sample_list)
    compound_num = len(sample_list[0])
    standard_dic = {}
    standard_dic_q = {}
    for i in range(4,compound_num):
        if sample_list[1][i] == "STD" or sample_list[1][i] == "standard":
                standard_dic[sample_list[2][i]] = [sample_list[j][i] for j in range(4, sample_num)]
    #creat a dictionary having stds
    for i in range(4, compound_num):
        standard_group = sample_list[2][i]
        if standard_group in standard_dic:
            for j in range(4, sample_num):
                if standard_dic[standard_group][j-4] != 0:
                    sample_list[j][i] /= standard_dic[standard_group][j-4]
                    sample_list[j][i] *= float(sample_list[3][i])
                else:
                    sample_list[j][i] = 0
        #fix by STD
        else:
            sample_list[2][i] = "not fixed"
    return

def tissue_weight_fix(sample_list):
    for i in range(4, len(sample_list)):
        for j in range(4, len(sample_list[0])):
            sample_list[i][j] /= sample_list[i][3]
    #fix by tissue weight
    return

def rearrange(sample_list):
    sample_num = len(sample_list)
    compound_num = len(sample_list[0])
    temp_list = []
    for i in range(sample_num):
            temp_list.append(sample_list[i][:4])
    #get by 3rd column
    for i in range(-(-(compound_num-4)//2)):
        for j in range(sample_num):
            temp_list[j].append(sample_list[j][i*2 + 4])
    for i in range((compound_num-4)//2):
        for j in range(sample_num):
            temp_list[j].append(sample_list[j][i*2 + 5])
    #get from 4th column
    for i in range(sample_num):
        for j in range(compound_num):
            sample_list[i][j] = temp_list[i][j]
    return 

def gather_data(sample_list):
    sample_num = len(sample_list)
    compound_num = len(sample_list[0])
    gathered_sample_list = [[""]]
    gathered_sample_list[0] += sample_list[0][4:]
    cal_list = ["intact"]*(sample_num-4)
    for i in range(4, sample_num):
        if cal_list[i-4] != "Calculated":
            temp_list = [[] for j in range(compound_num-3)]
            temp_list[0] = str(sample_list[i][0])
            for j in range(4, sample_num):
                if sample_list[j][0] == temp_list[0]:
                    for k in range(4, compound_num):
                        temp_list[k-3].append(sample_list[j][k])
                        cal_list[j-4] = "Calculated"
            gathered_sample_list.append(temp_list)
    #store the data of same name samples in each list
    return gathered_sample_list

def cal_average(gathered_sample_list):
    sample_num = len(gathered_sample_list)
    compound_num = len(gathered_sample_list[0])
    average_list = [[] for i in range(sample_num)]
    average_list[0] = list(gathered_sample_list[0])
    average_list[0][0] = "平均"
    for i in range(1, sample_num):
        average_list[i].append (gathered_sample_list[i][0])
    for i in range(1,sample_num):
        for j in range(1, compound_num):
            average_list[i].append(stx.mean(gathered_sample_list[i][j]))
    #calculate averages
    return average_list

def cal_SD(gathered_sample_list):
    sample_num = len(gathered_sample_list)
    compound_num = len(gathered_sample_list[0])
    SD_list = [[] for i in range(sample_num)]
    SD_list[0] = list(gathered_sample_list[0])
    SD_list[0][0] = "SD"
    for i in range(1,sample_num):
        SD_list[i].append(gathered_sample_list[i][0])
    for i in range(1,sample_num):
        for j in range(1, compound_num):
            if len(gathered_sample_list[i][j]) != 1:
                SD_list[i].append(stx.stdev(gathered_sample_list[i][j]))
            else:
                SD_list[i].append("error")
    #calculate SDs
    return SD_list

def write_data(x,y,list2D,ws,unit1):
    for i in range(len(list2D)):
        for j in range(len(list2D[0])):
            ws.cell(row = i + x +1, column = j + y +1, value = list2D[i][j])
    #paste the 2 dimension list on the cell(x,y) #consider A1 as (0,0)
    ws.cell(row = i + x + 2, column = y + len(list2D[0]), value = unit1)
    return

